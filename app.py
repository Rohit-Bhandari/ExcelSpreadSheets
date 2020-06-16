import openpyxl as xl

workbook = xl.load_workbook("students-result.xlsx")
sheet = workbook["Sheet1"]
print(sheet.cell(4,4).value)

for row in range(2, sheet.max_row + 1):
    physics = sheet.cell(row, 3).value
    maths = sheet.cell(row, 3).value
    history = sheet.cell(row, 3).value
    geography = sheet.cell(row, 3).value
    biology = sheet.cell(row, 3).value
    chemistry = sheet.cell(row, 3).value

    # print(physics, maths, history, geography, biology, chemistry)
    total_marks = physics + maths + history + geography + biology + chemistry
    total_marks_cell = sheet.cell(row, 9)
    total_marks_cell.value = total_marks

    total_percentages = round(total_marks / 6)
    total_percentages_cell = sheet.cell(row, 10)
    total_percentages_cell.value = total_percentages

workbook.save("students-result-final.xlsx")
print("Doc saved successfully")